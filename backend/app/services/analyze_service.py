import io
import json
from dataclasses import dataclass
from decimal import Decimal
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import httpx
from loguru import logger
from app.core.config import settings
from app.parsers.factory import parse_statement
from app.parsers.base import BaseParser, RawTransaction

# ── Category rules (no Ollama needed for these) ──────────────────────────────
MERCHANT_RULES: dict[str, str] = {
    "SWIGGY": "Food & Dining", "ZOMATO": "Food & Dining", "DOMINOS": "Food & Dining",
    "MCDONALDS": "Food & Dining", "KFC": "Food & Dining", "SUBWAY": "Food & Dining",
    "STARBUCKS": "Food & Dining", "DUNZO": "Food & Dining", "BLINKIT": "Food & Dining",
    "ZEPTO": "Food & Dining", "BIGBASKET": "Groceries", "DMART": "Groceries",
    "RELIANCE FRESH": "Groceries", "MORE SUPERMARKET": "Groceries",
    "AMAZON": "Shopping", "FLIPKART": "Shopping", "MYNTRA": "Shopping",
    "MEESHO": "Shopping", "AJIO": "Shopping", "NYKAA": "Shopping", "SNAPDEAL": "Shopping",
    "NETFLIX": "Subscriptions", "SPOTIFY": "Subscriptions", "YOUTUBE": "Subscriptions",
    "HOTSTAR": "Subscriptions", "PRIME": "Subscriptions", "ZEE5": "Subscriptions",
    "SONYLIV": "Subscriptions", "JIOCINEMA": "Subscriptions", "APPLE": "Subscriptions",
    "HPCL": "Fuel", "BPCL": "Fuel", "IOCL": "Fuel", "INDIAN OIL": "Fuel",
    "SHELL": "Fuel", "ESSAR": "Fuel",
    "AIRTEL": "Bills & Utilities", "JIO": "Bills & Utilities", "BSNL": "Bills & Utilities",
    "VODAFONE": "Bills & Utilities", "VI ": "Bills & Utilities",
    "BESCOM": "Bills & Utilities", "TATAPOWER": "Bills & Utilities",
    "ELECTRICITY": "Bills & Utilities", "WATER BOARD": "Bills & Utilities",
    "MAKEMYTRIP": "Travel", "IRCTC": "Travel", "GOIBIBO": "Travel",
    "OLA": "Travel", "UBER": "Travel", "RAPIDO": "Travel", "REDBUS": "Travel",
    "INDIGO": "Travel", "SPICEJET": "Travel", "AIR INDIA": "Travel",
    "APOLLO": "Healthcare", "MEDPLUS": "Healthcare", "PHARMEASY": "Healthcare",
    "NETMEDS": "Healthcare", "1MG": "Healthcare", "PRACTO": "Healthcare",
    "LIC": "Insurance", "LICI": "Insurance", "SBI LIFE": "Insurance",
    "HDFC LIFE": "Insurance", "ICICI PRU": "Insurance", "BAJAJ ALLIANZ": "Insurance",
    "ATM": "ATM Withdrawal", "CASH WD": "ATM Withdrawal",
    "BYJU": "Education", "UNACADEMY": "Education", "COURSERA": "Education",
    "UDEMY": "Education", "WHITEHAT": "Education",
    "ZERODHA": "Investments", "GROWW": "Investments", "UPSTOX": "Investments",
    "KUVERA": "Investments", "PAYTM MONEY": "Investments",
    "EMI": "EMI / Loans", "LOAN": "EMI / Loans", "BAJAJ FINANCE": "EMI / Loans",
    "CGST": "Taxes & Charges", "SGST": "Taxes & Charges", "IGST": "Taxes & Charges",
    "PETRO SURCHARGE": "Taxes & Charges", "SURCHARGE WAIVER": "Taxes & Charges",
    "FCY MARKUP": "Taxes & Charges", "MARKUP FEE": "Taxes & Charges",
    "LATE PAYMENT": "Taxes & Charges", "FINANCE CHARGE": "Taxes & Charges",
    "CONSOLIDATED": "Taxes & Charges",
}

VALID_CATEGORIES = [
    "Food & Dining", "Groceries", "Shopping", "Fuel", "Travel",
    "Bills & Utilities", "Entertainment", "Healthcare", "Education",
    "Investments", "EMI / Loans", "ATM Withdrawal", "Subscriptions",
    "Insurance", "Taxes & Charges", "Others",
]


@dataclass
class CategorizedTransaction:
    date: str
    raw_merchant: str
    normalized_merchant: str
    amount: Decimal
    transaction_type: str
    category: str
    description: str


def _rule_category(merchant: str) -> str | None:
    upper = merchant.upper()
    for keyword, cat in MERCHANT_RULES.items():
        if keyword in upper:
            return cat
    return None


async def _groq_category(merchant: str) -> str:
    prompt = f"""Classify this Indian merchant/transaction into ONE category from this list:
{json.dumps(VALID_CATEGORIES)}

Merchant: "{merchant}"

Reply with ONLY valid JSON: {{"category": "..."}}"""

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {settings.GROQ_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": settings.GROQ_MODEL,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0,
                    "max_tokens": 50,
                    "response_format": {"type": "json_object"},
                },
            )
            resp.raise_for_status()
            content = resp.json()["choices"][0]["message"]["content"]
            cat = json.loads(content).get("category", "Others")
            return cat if cat in VALID_CATEGORIES else "Others"
    except Exception as e:
        logger.warning(f"Groq categorization failed for '{merchant}': {e}")
        return "Others"


async def _ollama_category(merchant: str) -> str:
    prompt = f"""Classify this Indian merchant into ONE category from this list:
{json.dumps(VALID_CATEGORIES)}

Merchant: "{merchant}"

Reply with ONLY valid JSON: {{"category": "..."}}"""

    for attempt in range(settings.OLLAMA_MAX_RETRIES):
        model = settings.OLLAMA_PRIMARY_MODEL if attempt == 0 else settings.OLLAMA_FALLBACK_MODEL
        try:
            async with httpx.AsyncClient(timeout=settings.OLLAMA_TIMEOUT) as client:
                resp = await client.post(
                    f"{settings.OLLAMA_BASE_URL}/api/generate",
                    json={"model": model, "prompt": prompt, "stream": False, "format": "json"},
                )
                resp.raise_for_status()
                data = json.loads(resp.json().get("response", "{}"))
                cat = data.get("category", "Others")
                return cat if cat in VALID_CATEGORIES else "Others"
        except Exception as e:
            logger.warning(f"Ollama attempt {attempt + 1} failed: {e}")
    return "Others"


# In-memory cache so repeated merchants in the same upload don't hit the AI twice
_session_cache: dict[str, str] = {}


async def categorize(merchant: str) -> str:
    if merchant in _session_cache:
        return _session_cache[merchant]
    rule = _rule_category(merchant)
    if rule:
        _session_cache[merchant] = rule
        return rule
    if settings.GROQ_API_KEY:
        cat = await _groq_category(merchant)
    else:
        cat = await _ollama_category(merchant)
    _session_cache[merchant] = cat
    return cat


async def process_statement(file_path: str) -> list[CategorizedTransaction]:
    raw_transactions, bank_name = parse_statement(file_path)
    logger.info(f"Parsed {len(raw_transactions)} transactions from {bank_name}")

    results = []
    for raw in raw_transactions:
        normalized = BaseParser._normalize_merchant(raw.raw_merchant)
        category = await categorize(normalized)
        results.append(CategorizedTransaction(
            date=str(raw.date),
            raw_merchant=raw.raw_merchant,
            normalized_merchant=normalized,
            amount=raw.amount,
            transaction_type=raw.transaction_type,
            category=category,
            description=raw.description or "",
        ))
    return results


# ── Excel generation ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")


def _style_headers(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22


def _auto_width(ws) -> None:
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 55)


def _add_alternating_rows(ws, start_row: int, num_rows: int) -> None:
    for row in range(start_row, start_row + num_rows):
        if row % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = ALT_FILL


def build_excel(transactions: list[CategorizedTransaction]) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _sheet_transactions(writer, transactions)
        _sheet_category_summary(writer, transactions)
        _sheet_merchant_summary(writer, transactions)
        _sheet_monthly_summary(writer, transactions)

    output.seek(0)
    return output.read()


def _sheet_transactions(writer, txns: list[CategorizedTransaction]) -> None:
    rows = [{
        "Date": t.date,
        "Raw Merchant": t.raw_merchant,
        "Merchant": t.normalized_merchant,
        "Category": t.category,
        "Amount (₹)": float(t.amount),
        "Type": t.transaction_type.upper(),
        "Description": t.description,
    } for t in txns]

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Transactions", index=False)
    ws = writer.sheets["Transactions"]
    _style_headers(ws, len(df.columns))
    _add_alternating_rows(ws, 2, len(df))
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _sheet_category_summary(writer, txns: list[CategorizedTransaction]) -> None:
    debits = [t for t in txns if t.transaction_type == "debit"]
    total = sum(t.amount for t in debits) or Decimal("1")

    from collections import defaultdict
    cats: dict[str, dict] = defaultdict(lambda: {"total": Decimal("0"), "count": 0})
    for t in debits:
        cats[t.category]["total"] += t.amount
        cats[t.category]["count"] += 1

    rows = sorted([{
        "Category": cat,
        "Total Spent (₹)": float(v["total"]),
        "Transactions": v["count"],
        "% of Total": round(float(v["total"] / total * 100), 1),
    } for cat, v in cats.items()], key=lambda x: -x["Total Spent (₹)"])

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Category Summary", index=False)
    ws = writer.sheets["Category Summary"]
    _style_headers(ws, len(df.columns))
    _add_alternating_rows(ws, 2, len(df))
    _auto_width(ws)


def _sheet_merchant_summary(writer, txns: list[CategorizedTransaction]) -> None:
    from collections import defaultdict
    merchants: dict[str, dict] = defaultdict(lambda: {"total": Decimal("0"), "count": 0, "category": ""})
    for t in txns:
        if t.transaction_type == "debit":
            merchants[t.normalized_merchant]["total"] += t.amount
            merchants[t.normalized_merchant]["count"] += 1
            merchants[t.normalized_merchant]["category"] = t.category

    rows = sorted([{
        "Merchant": m,
        "Category": v["category"],
        "Total Spent (₹)": float(v["total"]),
        "Transactions": v["count"],
    } for m, v in merchants.items()], key=lambda x: -x["Total Spent (₹)"])

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Merchant Summary", index=False)
    ws = writer.sheets["Merchant Summary"]
    _style_headers(ws, len(df.columns))
    _add_alternating_rows(ws, 2, len(df))
    _auto_width(ws)


def _sheet_monthly_summary(writer, txns: list[CategorizedTransaction]) -> None:
    from collections import defaultdict
    months: dict[str, dict] = defaultdict(lambda: {"debit": Decimal("0"), "credit": Decimal("0"), "count": 0})
    for t in txns:
        month = t.date[:7]  # YYYY-MM
        if t.transaction_type == "debit":
            months[month]["debit"] += t.amount
        else:
            months[month]["credit"] += t.amount
        months[month]["count"] += 1

    rows = sorted([{
        "Month": m,
        "Total Spent (₹)": float(v["debit"]),
        "Total Credit (₹)": float(v["credit"]),
        "Net (₹)": float(v["credit"] - v["debit"]),
        "Transactions": v["count"],
    } for m, v in months.items()], key=lambda x: x["Month"])

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name="Monthly Summary", index=False)
    ws = writer.sheets["Monthly Summary"]
    _style_headers(ws, len(df.columns))
    _add_alternating_rows(ws, 2, len(df))
    _auto_width(ws)
